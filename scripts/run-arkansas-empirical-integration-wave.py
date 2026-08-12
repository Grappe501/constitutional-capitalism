#!/usr/bin/env python3
"""
CC-ARKANSAS-EMPIRICAL-INTEGRATION-WAVE-1.0 / UPD-132
Integrate registered voters, ADWS demand bridge, county HPSA, FDIC local capital
into County×Year layer + Living Systems Explorer + Counterexample Register.
Observation first. 43% honesty hold. No site lock. No false shortages.
"""
from __future__ import annotations

import csv
import json
import math
import statistics
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DL = ROOT / ".local" / "downloads" / "empirical-wave3"
PROJECT = ROOT / "data" / "project"
IMPORT = ROOT / "data" / "imports" / "arkansas-empirical-integration"
REPORTS = ROOT / "reports"

GENERATED_AT = "2026-08-12"
DECISION_ID = "CC-DEC-119"
UPDATE_ID = "UPD-132"
WAVE_SLICE = "CC-ARKANSAS-EMPIRICAL-INTEGRATION-WAVE-1.0"

COUNTIES = [
    {"fips": "05001", "name": "Arkansas County"},
    {"fips": "05073", "name": "Lafayette County"},
    {"fips": "05093", "name": "Mississippi County"},
    {"fips": "05107", "name": "Phillips County"},
    {"fips": "05129", "name": "Searcy County"},
    {"fips": "05141", "name": "Van Buren County"},
    {"fips": "05145", "name": "White County"},
]
FIPS = {c["fips"] for c in COUNTIES}
NAME = {c["fips"]: c["name"] for c in COUNTIES}


def write_json(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def obs(domain, metric, value, unit, definition, source_id, source_url, confidence, limitations, **extra):
    row = {
        "domain": domain,
        "metric": metric,
        "value": value,
        "unit": unit,
        "definition": definition,
        "source_id": source_id,
        "source_url": source_url,
        "confidence": confidence,
        "limitations": limitations,
    }
    row.update(extra)
    return row


def load_eavs() -> dict[str, dict]:
    path = DL / "eavs" / "2024" / "2024_EAVS_for_Public_Release_nolabel_V2.csv"
    out = {}
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            if row.get("State_Abbr") != "AR":
                continue
            code = (row.get("FIPSCode") or "").strip()
            if len(code) < 5:
                continue
            fips = code[:5]
            if fips not in FIPS:
                continue

            def num(k):
                v = row.get(k)
                if v in (None, "", "-99", "-88", "-77"):
                    return None
                try:
                    return int(float(v))
                except ValueError:
                    return None

            out[fips] = {
                "registered_total_a1a": num("A1a"),
                "registered_active_a1b": num("A1b"),
                "registered_inactive_a1c": num("A1c"),
                "ballots_cast_f1a": num("F1a"),
                "jurisdiction": row.get("Jurisdiction_Name"),
                "election_cycle": "2024_general",
            }
    return out


def load_fdic() -> dict[str, dict[str, dict]]:
    """fips -> year -> {branches, deposits_thousands}"""
    out: dict[str, dict[str, dict]] = defaultdict(dict)
    for year in ("2023", "2024", "2025"):
        path = DL / "fdic" / f"sod_ar_{year}.csv"
        agg = defaultdict(lambda: {"branches": 0, "deposits_thousands": 0.0})
        with path.open(encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                try:
                    cnty = int(float(row["CNTYNUMB"]))
                except (ValueError, KeyError):
                    continue
                fips = f"05{cnty:03d}"
                if fips not in FIPS:
                    continue
                agg[fips]["branches"] += 1
                try:
                    dep = float(row.get("DEPSUMBR") or 0)
                except ValueError:
                    dep = 0
                agg[fips]["deposits_thousands"] += dep
        for fips, vals in agg.items():
            out[fips][year] = {
                "fdic_branch_count": vals["branches"],
                "fdic_deposits_usd": int(vals["deposits_thousands"] * 1000),
            }
    return out


def load_hpsa() -> dict[str, dict]:
    path = DL / "hrsa" / "BCD_HPSA_FCT_DET_PC.csv"
    out: dict[str, dict] = {
        f: {
            "designated_rows": 0,
            "max_hpsa_score": None,
            "geographic_designated": 0,
            "population_designated": 0,
            "facility_designated": 0,
            "fte_sum_designated": 0.0,
            "types": set(),
        }
        for f in FIPS
    }
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            if (row.get("Primary State Abbreviation") or row.get("State Abbreviation")) != "AR":
                continue
            if (row.get("HPSA Status") or "") != "Designated":
                continue
            fips = (row.get("Common State County FIPS Code") or "").zfill(5)
            if fips not in FIPS:
                # try geography id
                geo = (row.get("HPSA Geography Identification Number") or "").zfill(5)
                if geo in FIPS:
                    fips = geo
                else:
                    continue
            dtype = row.get("Designation Type") or ""
            out[fips]["designated_rows"] += 1
            out[fips]["types"].add(dtype)
            try:
                score = float(row.get("HPSA Score") or 0)
            except ValueError:
                score = None
            if score is not None:
                cur = out[fips]["max_hpsa_score"]
                out[fips]["max_hpsa_score"] = score if cur is None else max(cur, score)
            try:
                fte = float(row.get("HPSA FTE") or 0)
            except ValueError:
                fte = 0
            out[fips]["fte_sum_designated"] += fte
            dl = dtype.lower()
            if "geographic" in dl:
                out[fips]["geographic_designated"] += 1
            elif "population" in dl:
                out[fips]["population_designated"] += 1
            else:
                out[fips]["facility_designated"] += 1
    for f in out:
        out[f]["types"] = sorted(out[f]["types"])
        out[f]["primary_care_hpsa_designated"] = 1 if out[f]["designated_rows"] else 0
    return out


def load_adws_bridge(cip_bind: dict) -> dict:
    # SOC-CIP crosswalk
    wb = openpyxl.load_workbook(DL / "adws" / "SOC-CIP-Crosswalk.xlsx", data_only=True)
    ws = wb["SOC-CIP"]
    cip_to_socs: dict[str, set[str]] = defaultdict(set)
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[0] or not row[2]:
            continue
        soc = str(row[0]).strip()
        cip = str(row[2]).strip().rstrip(".")
        if len(cip.split(".")) == 2:
            # normalize 44.0401
            left, right = cip.split(".")
            cip = f"{left.zfill(2)}.{right[:4].ljust(4, '0')[:4]}"
        cip_to_socs[cip].add(soc)
        cip_to_socs[cip[:2]].add(soc)

    # Occupations demand
    wb2 = openpyxl.load_workbook(DL / "adws" / "LT-Proj-State-20-30.xlsx", data_only=True)
    ws2 = wb2["Occupation"]
    soc_demand = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        soc = str(row[0]).strip()
        if soc.endswith("0000") or soc.endswith("-0000"):
            continue  # major groups optional; keep detailed
        try:
            emp2020 = int(float(row[2])) if row[2] is not None else None
            emp2030 = int(float(row[3])) if row[3] is not None else None
            openings = int(float(row[9])) if row[9] is not None else None
        except (ValueError, TypeError):
            continue
        # Keep 6-digit SOC (xx-xxxx)
        if len(soc) >= 7 and soc[2] == "-":
            soc_demand[soc] = {
                "soc": soc,
                "title": row[1],
                "employment_2020": emp2020,
                "employment_2030": emp2030,
                "total_annual_openings": openings,
            }

    # Aggregate 2023 IPEDS completions by CIP
    cip_comp = defaultdict(int)
    for rec in cip_bind.get("records", []):
        if rec.get("year") != "2023":
            continue
        cip_comp[rec["cipcode"]] += rec["completions"]

    # Bridge family-level: map workforce categories via CIP2 → SOCs → openings
    family_to_cip2 = {
        "agriculture_food": ["01", "03"],
        "computer_ai_tech": ["11"],
        "engineering_advanced_mfg": ["14", "15"],
        "education_human_services": ["13", "44", "19"],
        "healthcare": ["51", "60"],
        "trades_transport_logistics": ["46", "47", "48", "49"],
        "business_management": ["52"],
        "public_law": ["22", "43"],
    }

    family_bridge = []
    for family, cip2s in family_to_cip2.items():
        completions = sum(v for c, v in cip_comp.items() if c[:2] in cip2s)
        socs = set()
        for c2 in cip2s:
            socs |= cip_to_socs.get(c2, set())
        # also exact CIPs
        for cip, sc in cip_to_socs.items():
            if "." in cip and cip[:2] in cip2s:
                socs |= sc
        openings = 0
        matched_socs = 0
        for soc in socs:
            if soc in soc_demand and soc_demand[soc]["total_annual_openings"] is not None:
                openings += soc_demand[soc]["total_annual_openings"]
                matched_socs += 1
        ratio = round(completions / openings, 3) if openings else None
        family_bridge.append(
            {
                "workforce_category": family,
                "geography": "arkansas_statewide",
                "ipeds_completions_2023": completions,
                "adws_linked_soc_count": matched_socs,
                "adws_total_annual_openings_sum": openings,
                "capacity_to_openings_ratio": ratio,
                "chain": "training_capacity → graduates(IPEDS) → occupations(SOC-CIP) → demand(ADWS statewide) → geographic_gap(NEE_COUNTY)",
                "not_a_shortage_claim": True,
                "notes": [
                    "Low completions alone ≠ shortage",
                    "Demand is statewide/LWDA — not county FIPS",
                    "Ratio is descriptive bridge, not adequacy verdict",
                ],
            }
        )

    return {
        "version": "1.0.0",
        "slice_id": "CC-ARKANSAS-ADWS-DEMAND-CIP-JOIN-1.0",
        "wave_slice_id": WAVE_SLICE,
        "generated_at": GENERATED_AT,
        "decision_id": DECISION_ID,
        "update_id": UPDATE_ID,
        "status": "PASSED_STATEWIDE_BRIDGE_COUNTY_GAP_NEE",
        "purpose": "Join educational capacity to ADWS occupational demand without inventing county shortages.",
        "distinction": {
            "educational_capacity": "IPEDS completions by CIP",
            "employment_demand": "ADWS long-term occupational openings (statewide)",
            "geographic_gap": "NEE at county — ADWS files are state/LWDA",
        },
        "sources": [
            {
                "id": "ADWS-SOC-CIP",
                "url": "https://www.discover.arkansas.gov/_docs/Publications/Projected-Employment-Opportunities-List1/SOC-CIP-Crosswalk.xlsx",
            },
            {
                "id": "ADWS-LT-20-30",
                "url": "https://www.discover.arkansas.gov/_docs/Publications/Projections/2020-2030/LT-Proj-State-20-30.xlsx",
            },
            {"id": "IPEDS-C2023", "path": "data/project/arkansas_cip_institution_year_completer_bind.json"},
        ],
        "family_bridge": sorted(family_bridge, key=lambda x: x["workforce_category"]),
        "soc_demand_rows": len(soc_demand),
        "cip_soc_links": sum(len(v) for v in cip_to_socs.values()),
        "holds": [
            "Do not call a shortage merely because completions are low",
            "Statewide demand ≠ county gap",
            "Observation first",
        ],
    }


def latest_metric(county, metric):
    for yr in reversed(county["years"]):
        for o in yr["observations"]:
            if o.get("metric") == metric and o.get("value") is not None:
                return o["value"], yr["year"]
    return None, None


def expand_layer(layer, eavs, fdic, hpsa) -> tuple[dict, int]:
    layer = json.loads(json.dumps(layer))
    layer["version"] = "1.2.0"
    layer["slice_id"] = "CC-ARKANSAS-COUNTY-LONGITUDINAL-OBSERVATION-LAYER-1.2"
    layer["wave_slice_id"] = WAVE_SLICE
    layer["decision_id"] = DECISION_ID
    layer["update_id"] = UPDATE_ID
    layer["status"] = "LAYER_V1_2_INTEGRATION"
    layer["generated_at"] = GENERATED_AT
    layer["research_question"] = (
        "Do communities with stronger institutional capacity, economic opportunity, healthcare access, "
        "educational pathways, local capital, and civic participation exhibit different trajectories over time?"
    )
    layer["analytical_mode"] = [
        "which_variables_move_together",
        "which_move_first",
        "which_do_not_move_together",
        "where_are_counterexamples",
        "which_relationships_survive_cross_county_comparison",
    ]
    county_map = {c["fips"]: c for c in layer["counties"]}
    added = 0

    def ensure_year(fips, year):
        c = county_map[fips]
        for yr in c["years"]:
            if yr["year"] == year:
                return yr
        yr = {"year": year, "observations": []}
        c["years"].append(yr)
        c["years"].sort(key=lambda x: x["year"])
        return yr

    def has(yr, metric, source_id):
        return any(
            o.get("metric") == metric and o.get("source_id") == source_id and o.get("value") is not None
            for o in yr["observations"]
        )

    # EAVS 2024 registered + turnout of registered for 2024 presidential
    for fips, vals in eavs.items():
        yr = ensure_year(fips, "2024")
        if vals["registered_total_a1a"] is not None and not has(yr, "registered_voters_total", "EAC-EAVS-2024"):
            yr["observations"].append(
                obs(
                    "civic",
                    "registered_voters_total",
                    vals["registered_total_a1a"],
                    "persons",
                    "EAVS A1a total registered voters (2024 cycle jurisdiction report)",
                    "EAC-EAVS-2024",
                    "https://www.eac.gov/research-and-data/studies-and-reports",
                    "verified_primary",
                    ["Point-in-time registration for EAVS 2024 cycle — not a multi-year series yet"],
                    election_type="registration_stock",
                )
            )
            yr["observations"].append(
                obs(
                    "civic",
                    "registered_voters_active",
                    vals["registered_active_a1b"],
                    "persons",
                    "EAVS A1b active registered voters",
                    "EAC-EAVS-2024",
                    "https://www.eac.gov/research-and-data/studies-and-reports",
                    "verified_primary",
                    [],
                    election_type="registration_stock",
                )
            )
            added += 2
            if vals["ballots_cast_f1a"] is not None:
                yr["observations"].append(
                    obs(
                        "civic",
                        "ballots_cast_eavs",
                        vals["ballots_cast_f1a"],
                        "ballots",
                        "EAVS F1a total ballots cast (2024 general cycle)",
                        "EAC-EAVS-2024",
                        "https://www.eac.gov/research-and-data/studies-and-reports",
                        "verified_primary",
                        ["Election-type: 2024 general; preserve distinction from presidential-only totals"],
                        election_type="general_2024",
                    )
                )
                added += 1
                if vals["registered_total_a1a"]:
                    yr["observations"].append(
                        obs(
                            "civic",
                            "general_turnout_of_registered_percent",
                            round(100.0 * vals["ballots_cast_f1a"] / vals["registered_total_a1a"], 1),
                            "percent",
                            "EAVS ballots cast / total registered × 100 (2024 general)",
                            "DERIVED-EAVS-TURNOUT",
                            "https://www.eac.gov/research-and-data/studies-and-reports",
                            "derived",
                            [
                                "Registered-voter turnout — distinct from CVAP turnout",
                                "Preserve election-type labels",
                            ],
                            election_type="general_2024",
                        )
                    )
                    added += 1
            # Presidential votes / registered (if presidential totals present)
            pres = next(
                (
                    o["value"]
                    for o in yr["observations"]
                    if o.get("metric") == "presidential_total_votes" and o.get("value") is not None
                ),
                None,
            )
            if pres and vals["registered_total_a1a"]:
                yr["observations"].append(
                    obs(
                        "civic",
                        "presidential_turnout_of_registered_percent",
                        round(100.0 * pres / vals["registered_total_a1a"], 1),
                        "percent",
                        "Presidential total votes / EAVS registered total × 100",
                        "DERIVED-PRES-EAVS",
                        "https://www.eac.gov/research-and-data/studies-and-reports",
                        "derived",
                        [
                            "Uses 2024 EAVS registration stock with 2024 presidential votes",
                            "Distinct from general_turnout_of_registered_percent and CVAP turnout",
                        ],
                        election_type="presidential_2024",
                    )
                )
                added += 1

    # FDIC
    for fips, years in fdic.items():
        for year, vals in years.items():
            yr = ensure_year(fips, year)
            if not has(yr, "fdic_branch_count", "FDIC-SOD"):
                yr["observations"].append(
                    obs(
                        "local_capital",
                        "fdic_branch_count",
                        vals["fdic_branch_count"],
                        "branches",
                        "Count of FDIC SOD branch records in county",
                        "FDIC-SOD",
                        "https://banks.data.fdic.gov/bankfind-suite/SOD",
                        "verified_primary",
                        ["Branch presence ≠ community prosperity accounts", "Deposits are SOD branch deposits"],
                    )
                )
                yr["observations"].append(
                    obs(
                        "local_capital",
                        "fdic_deposits_usd",
                        vals["fdic_deposits_usd"],
                        "usd",
                        "Sum of FDIC SOD DEPSUMBR (converted from $ thousands) for county branches",
                        "FDIC-SOD",
                        "https://banks.data.fdic.gov/bankfind-suite/SOD",
                        "verified_primary",
                        ["Deposit concentration ≠ local ownership", "Headquarters vs branch deposits can mislead"],
                    )
                )
                added += 2

    # HPSA — attach to 2024/2025 as current designation snapshot
    for fips, vals in hpsa.items():
        yr = ensure_year(fips, "2025")
        if not has(yr, "primary_care_hpsa_designated", "HRSA-HPSA-PC"):
            yr["observations"].append(
                obs(
                    "healthcare",
                    "primary_care_hpsa_designated",
                    vals["primary_care_hpsa_designated"],
                    "flag",
                    "1 if any Designated HRSA primary-care HPSA row intersects county",
                    "HRSA-HPSA-PC",
                    "https://data.hrsa.gov/data/download",
                    "verified_primary",
                    [
                        "Overlapping geographic/population/facility designations — do not sum to population share without method lock",
                        "Snapshot of download date designations",
                    ],
                )
            )
            if vals["max_hpsa_score"] is not None:
                yr["observations"].append(
                    obs(
                        "healthcare",
                        "primary_care_hpsa_max_score",
                        vals["max_hpsa_score"],
                        "score_0_25",
                        "Maximum HPSA score among Designated primary-care rows in county",
                        "HRSA-HPSA-PC",
                        "https://data.hrsa.gov/data/download",
                        "verified_primary",
                        ["Higher score ≈ greater shortage intensity in HRSA methodology"],
                    )
                )
                added += 1
            yr["observations"].append(
                obs(
                    "healthcare",
                    "primary_care_hpsa_designated_row_count",
                    vals["designated_rows"],
                    "count",
                    "Count of Designated primary-care HPSA detail rows intersecting county",
                    "HRSA-HPSA-PC",
                    "https://data.hrsa.gov/data/download",
                    "verified_primary",
                    ["Row count ≠ unique population covered"],
                )
            )
            added += 2

    # Deposits per capita where PEP pop available
    for fips, c in county_map.items():
        pop, _ = latest_metric(c, "population_total")
        dep, depy = latest_metric(c, "fdic_deposits_usd")
        if pop and dep:
            yr = ensure_year(fips, depy or "2025")
            if not has(yr, "fdic_deposits_per_capita_usd", "DERIVED-FDIC-PEP"):
                yr["observations"].append(
                    obs(
                        "local_capital",
                        "fdic_deposits_per_capita_usd",
                        round(dep / pop, 0),
                        "usd_per_person",
                        "FDIC SOD county deposits / nearest PEP population",
                        "DERIVED-FDIC-PEP",
                        "https://banks.data.fdic.gov/bankfind-suite/SOD",
                        "derived",
                        ["Crude intensity proxy — not wealth or ownership"],
                    )
                )
                added += 1

    non_null = 0
    for c in layer["counties"]:
        for yr in c["years"]:
            non_null += sum(1 for o in yr["observations"] if o.get("value") is not None)
    layer["stats"] = {
        "county_count": len(layer["counties"]),
        "year_span": [
            min(y["year"] for c in layer["counties"] for y in c["years"]),
            max(y["year"] for c in layer["counties"] for y in c["years"]),
        ],
        "non_null_observations": non_null,
        "observations_added_this_wave": added,
    }
    layer["blockers"] = [
        {"id": "registered_voter_multi_year_nee", "detail": "Only EAVS 2024 bound; SOS historical PDF series still NEE"},
        {"id": "adws_county_demand_nee", "detail": "ADWS demand is statewide/LWDA — county geographic gap NEE"},
        {"id": "maternal_hospital_closure_nee", "detail": "County maternal/hospital closure series not bound this pass"},
        {"id": "ownership_density_nee", "detail": "Local business/ownership density still NEE beyond FDIC/NASS"},
    ]
    return layer, added


def zscore_map(values: dict[str, float]) -> dict[str, float]:
    xs = list(values.values())
    if len(xs) < 2:
        return {k: 0.0 for k in values}
    mu = statistics.mean(xs)
    sd = statistics.pstdev(xs) or 1.0
    return {k: (v - mu) / sd for k, v in values.items()}


def build_counterexample_register(layer: dict) -> dict:
    """For major co-travel hypotheses, find supporting / contradictory / ambiguous / missing cases."""
    counties = layer["counties"]

    def series(metric):
        out = {}
        for c in counties:
            v, y = latest_metric(c, metric)
            if v is not None:
                out[c["fips"]] = {"value": float(v), "year": y, "county": c["county"]}
        return out

    # Prefer richer metrics when present
    civic = series("presidential_turnout_of_registered_percent") or series(
        "presidential_turnout_of_cvap_percent"
    ) or series("presidential_votes_per_1000_population")
    capital = series("fdic_deposits_per_capita_usd") or series("fdic_deposits_usd")
    poverty = series("acs5_poverty_rate_all_ages") or series("poverty_rate_all_ages")
    farms = series("farm_operations")
    hpsa = series("primary_care_hpsa_max_score")
    income = series("acs5_median_household_income") or series("median_household_income")

    hypotheses = []

    def add_hyp(hyp_id, claim, x_name, x, y_name, y, expected="positive"):
        xz = zscore_map({k: v["value"] for k, v in x.items()})
        yz = zscore_map({k: v["value"] for k, v in y.items()})
        common = sorted(set(xz) & set(yz))
        if len(common) < 3:
            hypotheses.append(
                {
                    "hypothesis_id": hyp_id,
                    "claim": claim,
                    "status": "MISSING_DATA",
                    "x": x_name,
                    "y": y_name,
                    "note": "Fewer than 3 counties with both metrics",
                }
            )
            return
        # product of z for expected positive co-travel; for negative expected, flip y
        scored = []
        for f in common:
            yz_adj = yz[f] if expected == "positive" else -yz[f]
            scored.append(
                {
                    "fips": f,
                    "county": NAME[f],
                    "x_z": round(xz[f], 3),
                    "y_z": round(yz[f], 3),
                    "support_score": round(xz[f] * yz_adj, 3),
                    "x_value": x[f]["value"],
                    "y_value": y[f]["value"],
                    "x_year": x[f]["year"],
                    "y_year": y[f]["year"],
                }
            )
        scored.sort(key=lambda r: r["support_score"], reverse=True)
        # Counterexamples: high x + low y or low x + high y relative to median
        med_x = statistics.median([xz[f] for f in common])
        med_y = statistics.median([yz[f] for f in common])
        contradictions = []
        for f in common:
            if expected == "positive":
                if xz[f] > med_x and yz[f] < med_y:
                    contradictions.append(
                        {
                            "pattern": "strong_x_weak_y",
                            "fips": f,
                            "county": NAME[f],
                            "x_z": round(xz[f], 3),
                            "y_z": round(yz[f], 3),
                            "why_valuable": "Economic/institutional strength without matching civic participation — theory gap candidate",
                        }
                    )
                if xz[f] < med_x and yz[f] > med_y:
                    contradictions.append(
                        {
                            "pattern": "weak_x_strong_y",
                            "fips": f,
                            "county": NAME[f],
                            "x_z": round(xz[f], 3),
                            "y_z": round(yz[f], 3),
                            "why_valuable": "Strong participation despite weaker structural conditions — theory gap candidate",
                        }
                    )
        hypotheses.append(
            {
                "hypothesis_id": hyp_id,
                "claim": claim,
                "expected_co_travel": expected,
                "x": x_name,
                "y": y_name,
                "n_counties": len(common),
                "strongest_supporting_pattern": scored[0],
                "strongest_contradictory_pattern": scored[-1],
                "counterexamples": contradictions,
                "ambiguous_middle": [s for s in scored if abs(s["support_score"]) < 0.25][:3],
                "definition_breaks": [
                    "Cross-sectional z-scores on latest available years — not causal",
                    "Metric years may differ across series",
                    "Seven-county designated set only — not statewide inference",
                ],
                "missing_data_cases": [
                    NAME[f] for f in FIPS if f not in common
                ],
            }
        )

    if capital and civic:
        add_hyp(
            "HYP-CECS-LOCAL-CAPITAL-CIVIC",
            "Stronger local capital intensity tends to co-travel with stronger civic participation",
            "fdic_deposits_per_capita_or_deposits",
            capital,
            "presidential_turnout_registered_or_cvap",
            civic,
            "positive",
        )
    if poverty and civic:
        add_hyp(
            "HYP-CECS-POVERTY-CIVIC",
            "Higher poverty tends to co-travel with weaker civic participation (descriptive test — not destiny)",
            "poverty_rate",
            poverty,
            "presidential_turnout_registered_or_cvap",
            civic,
            "negative",
        )
    if farms and civic:
        add_hyp(
            "HYP-CECS-FARM-STRUCTURE-CIVIC",
            "Denser farm-operation structure tends to co-travel with civic participation patterns",
            "farm_operations",
            farms,
            "presidential_turnout_registered_or_cvap",
            civic,
            "positive",
        )
    if hpsa and poverty:
        add_hyp(
            "HYP-CECS-HPSA-POVERTY",
            "Primary-care shortage intensity tends to co-travel with poverty",
            "primary_care_hpsa_max_score",
            hpsa,
            "poverty_rate",
            poverty,
            "positive",
        )
    if income and capital:
        add_hyp(
            "HYP-CECS-INCOME-LOCAL-CAPITAL",
            "Higher median income tends to co-travel with higher local deposit intensity",
            "median_household_income",
            income,
            "fdic_deposits_per_capita_or_deposits",
            capital,
            "positive",
        )

    return {
        "version": "1.0.0",
        "slice_id": "CC-ARKANSAS-COUNTY-COUNTEREXAMPLE-REGISTER-1.0",
        "wave_slice_id": WAVE_SLICE,
        "generated_at": GENERATED_AT,
        "decision_id": DECISION_ID,
        "update_id": UPDATE_ID,
        "status": "REGISTER_V1_DESIGNATED_SET",
        "purpose": "Protect against confirmation machines: for each major co-travel hypothesis, surface strongest support, strongest contradiction, ambiguous and missing-data cases.",
        "method_wall": [
            "observation_first",
            "interpretation_second",
            "causation_only_after_modeling",
        ],
        "not": [
            "causal_tests",
            "public_scoreboard",
            "hypothesis_promotion",
            "statewide_inference_from_7_counties",
        ],
        "hypotheses": hypotheses,
        "note": "Counterexamples are not nuisances — they reveal what the theory is missing.",
    }


def rebuild_explorer(layer, cip, adws, counter) -> dict:
    prior = json.loads(
        (PROJECT / "arkansas_county_living_systems_explorer.json").read_text(encoding="utf-8")
    )
    counties = []
    for c in layer["counties"]:
        timeline = []
        for yr in c["years"]:
            metrics = {
                o["metric"]: {
                    "value": o["value"],
                    "unit": o.get("unit"),
                    "source_id": o.get("source_id"),
                    "definition": o.get("definition"),
                    "election_type": o.get("election_type"),
                }
                for o in yr["observations"]
                if o.get("value") is not None
            }
            if metrics:
                timeline.append({"year": yr["year"], "metrics": metrics})
        counties.append(
            {
                "fips": c["fips"],
                "county": c["county"],
                "role": c.get("role"),
                "timeline": timeline,
            }
        )
    compare_metrics = sorted(
        set(prior.get("compare_metrics", []))
        | {
            "registered_voters_total",
            "general_turnout_of_registered_percent",
            "presidential_turnout_of_registered_percent",
            "presidential_turnout_of_cvap_percent",
            "fdic_branch_count",
            "fdic_deposits_usd",
            "fdic_deposits_per_capita_usd",
            "primary_care_hpsa_designated",
            "primary_care_hpsa_max_score",
            "gubernatorial_total_votes",
            "acs5_unemployment_rate",
            "farm_operations",
        }
    )
    return {
        "version": "1.1.0",
        "slice_id": "CC-ARKANSAS-COUNTY-LIVING-SYSTEMS-EXPLORER-1.1",
        "wave_slice_id": WAVE_SLICE,
        "generated_at": GENERATED_AT,
        "decision_id": DECISION_ID,
        "update_id": UPDATE_ID,
        "status": "INTERNAL_INTEGRATED_LAB_V1",
        "not": [
            "public_scoreboard",
            "community_health_index",
            "causal_model",
            "locked_pilot_ranking",
        ],
        "method_wall": [
            "observation_first",
            "interpretation_second",
            "causation_only_after_modeling",
        ],
        "research_questions": [
            "Do communities with stronger institutional capacity, economic opportunity, healthcare access, educational pathways, local capital, and civic participation exhibit different trajectories over time?",
            "Which variables move together?",
            "Which move first?",
            "Which don't move together?",
            "Where are the counterexamples?",
            "Which relationships survive comparison across counties?",
        ],
        "compare_metrics": compare_metrics,
        "counties": counties,
        "education_workforce_bridge": {
            "status": cip.get("status"),
            "completions_by_workforce_category": cip.get("completions_by_workforce_category"),
            "keystone_capacity_signals": cip.get("keystone_capacity_signals"),
            "distinction": cip.get("distinction"),
            "adws_state_bridge": adws.get("family_bridge"),
            "adws_status": adws.get("status"),
        },
        "counterexample_register_path": "data/project/arkansas_county_counterexample_register.json",
        "counterexample_register": {
            "hypothesis_count": len(counter.get("hypotheses", [])),
            "hypotheses": counter.get("hypotheses", []),
        },
        "candidates_not_locked": ["AR-GEO-ROSE-BUD", "AR-GEO-LEWISVILLE"],
        "source_layer": "data/project/arkansas_county_longitudinal_observation_layer.json",
        "dial_note": "Overall completion remains 43% — richer evidence infrastructure does not inflate the dial",
    }


def build_wave(layer, adws, counter, explorer, added) -> dict:
    return {
        "version": "1.0.0",
        "slice_id": WAVE_SLICE,
        "generated_at": GENERATED_AT,
        "status": "wave_executed_integration_partial",
        "module_id": "CC-MOD-ARKANSAS-EMPIRICAL-INTEGRATION-WAVE",
        "decision_id": DECISION_ID,
        "update_id": UPDATE_ID,
        "amends_decision": "CC-DEC-118",
        "signal": "Integration, not architecture: registered voters, ADWS capacity↔demand bridge, county HPSA, FDIC local capital → Living Systems Explorer + Counterexample Register.",
        "method_wall": [
            "observation_first",
            "interpretation_second",
            "causation_only_after_modeling",
        ],
        "progression_lock": [
            "evidence_infrastructure",
            "validated_comparative_diagnosis",
            "actual_models",
            "legal_feasibility",
            "pilot_design",
            "measured_intervention",
        ],
        "dependency_chain": [
            {"pass": 1, "id": "registered_voters_eavs", "status": "PASSED_EAVS_2024"},
            {"pass": 2, "id": "adws_demand_cip_bridge", "status": adws["status"], "artifact": "data/project/arkansas_adws_demand_cip_bridge.json"},
            {"pass": 3, "id": "county_hpsa", "status": "PASSED_HRSA_PC_DESIGNATED"},
            {"pass": 4, "id": "fdic_local_capital", "status": "PASSED_SOD_2023_2025"},
            {"pass": 5, "id": "counterexample_register", "status": counter["status"], "artifact": "data/project/arkansas_county_counterexample_register.json"},
            {"pass": 6, "id": "explorer_integration", "status": explorer["status"], "board_path": "/county-living-systems/"},
        ],
        "completion_rule": {
            "overall_percent_held": 43,
            "dial_meaning": "Do not inflate dial because evidence infrastructure got richer. Next genuine increase after validated comparative diagnosis → models → legal feasibility → pilot → measured intervention.",
        },
        "holds": [
            "No locked pilot site",
            "No public score",
            "No shortage claim from low completions alone",
            "ADWS demand statewide ≠ county gap",
            "43% honesty hold",
            "Counterexamples privileged",
        ],
        "structural_active_unchanged": "CC-PHASE-2.1-BASELINE-DEFINITION-LOCKS-AND-NEXT-LEGITIMATE-FILLS-1.0",
        "return": "reports/CC_ARKANSAS_EMPIRICAL_INTEGRATION_WAVE_1_0_RETURN.md",
        "stats": {
            "longitudinal_non_null_observations": layer["stats"]["non_null_observations"],
            "observations_added": added,
            "counterexample_hypotheses": len(counter.get("hypotheses", [])),
        },
    }


def write_return(wave, layer, adws, counter) -> str:
    lines = [
        "# CC-ARKANSAS-EMPIRICAL-INTEGRATION-WAVE-1.0 — Return",
        "",
        f"**Slice ID:** `{WAVE_SLICE}`  ",
        "**Status:** PASSED (integration partial)  ",
        f"**Date:** {GENERATED_AT}  ",
        f"**Decision:** `{DECISION_ID}`  ",
        f"**Update:** `{UPDATE_ID}`",
        "",
        "## Signal",
        "",
        "Integration, not architecture. The County Living Systems Explorer becomes the project's first integrated empirical laboratory.",
        "",
        "## Method wall",
        "",
        "**Observation first → interpretation second → causation only after modeling.**",
        "",
        "Ask which variables move together / first / not together — and where counterexamples live. Not yet “does X cause Y?”",
        "",
        "## Four joins",
        "",
        "1. **Registered voters** — EAVS 2024 A1a/A1b + F1a; general and presidential turnout-of-registered; election-type labels preserved; multi-year SOS series still NEE.",
        "2. **Workforce demand** — ADWS SOC–CIP + LT 2020–2030 openings joined to IPEDS 2023 completions at statewide family level; **geographic county gap NEE**; no shortage-from-low-completions claims.",
        "3. **Healthcare** — HRSA primary-care Designated HPSA rows for all 7 counties (score + row counts).",
        "4. **Local capital** — FDIC SOD branches/deposits 2023–2025 + deposits/capita; farm structure already present.",
        "",
        f"County×Year non-null observations: **{layer['stats']['non_null_observations']}** (+{layer['stats']['observations_added_this_wave']}).",
        "",
        "## Counterexample Register",
        "",
        f"- Hypotheses tested descriptively: **{len(counter.get('hypotheses', []))}**",
        "- Each entry surfaces strongest support, strongest contradiction, ambiguous middle, missing-data, definition breaks.",
        "- Explicitly hunts **strong economy / weak participation** and **weak economy / strong participation** patterns.",
        "",
    ]
    for h in counter.get("hypotheses", []):
        if h.get("status") == "MISSING_DATA":
            lines.append(f"- `{h['hypothesis_id']}`: missing data")
            continue
        supp = h.get("strongest_supporting_pattern") or {}
        contra = h.get("counterexamples") or []
        lines.append(
            f"- `{h['hypothesis_id']}`: support {supp.get('county')} (score {supp.get('support_score')}); "
            f"counterexamples {len(contra)}"
        )
    lines += [
        "",
        "## Holds",
        "",
        "- Overall dial **43%** (do not inflate for richer infrastructure)",
        "- Rose Bud / Lewisville **not locked**",
        "- No public score / CHI",
        "- Capacity ≠ demand; statewide ADWS ≠ county shortage",
        "",
        "## Progression (unchanged)",
        "",
        "Evidence infrastructure → validated comparative diagnosis → models → legal feasibility → pilot design → measured intervention",
        "",
        "## Next",
        "",
        "1. Multi-year registered-voter / SOS series",
        "2. LWDA→county demand allocation method (or refuse)",
        "3. Maternal/hospital closure series where official",
        "4. Begin validated comparative diagnosis packets using Counterexample Register",
        "",
    ]
    return "\n".join(lines)


def main():
    layer = json.loads(
        (PROJECT / "arkansas_county_longitudinal_observation_layer.json").read_text(encoding="utf-8")
    )
    cip = json.loads(
        (PROJECT / "arkansas_cip_institution_year_completer_bind.json").read_text(encoding="utf-8")
    )
    print("loading eavs/fdic/hpsa/adws...")
    eavs = load_eavs()
    fdic = load_fdic()
    hpsa = load_hpsa()
    print(f"eavs={len(eavs)} fdic_counties={len(fdic)} hpsa={sum(1 for v in hpsa.values() if v['designated_rows'])}")
    adws = load_adws_bridge(cip)
    layer, added = expand_layer(layer, eavs, fdic, hpsa)
    counter = build_counterexample_register(layer)
    explorer = rebuild_explorer(layer, cip, adws, counter)
    wave = build_wave(layer, adws, counter, explorer, added)
    ret = write_return(wave, layer, adws, counter)

    write_json(PROJECT / "arkansas_county_longitudinal_observation_layer.json", layer)
    write_json(PROJECT / "arkansas_adws_demand_cip_bridge.json", adws)
    write_json(PROJECT / "arkansas_county_counterexample_register.json", counter)
    write_json(PROJECT / "arkansas_county_living_systems_explorer.json", explorer)
    write_json(PROJECT / "arkansas_empirical_integration_wave.json", wave)
    (REPORTS / "CC_ARKANSAS_EMPIRICAL_INTEGRATION_WAVE_1_0_RETURN.md").write_text(ret, encoding="utf-8")

    IMPORT.mkdir(parents=True, exist_ok=True)
    write_json(IMPORT / "manifest.json", {
        "import_id": "arkansas-empirical-integration-1.0",
        "generated_at": GENERATED_AT,
        "decision_id": DECISION_ID,
        "contains_api_keys": False,
        "sources": ["EAC EAVS 2024", "HRSA HPSA PC", "FDIC SOD", "ADWS SOC-CIP + LT projections"],
    })
    write_json(IMPORT / "designated_county_eavs_2024.json", {"rows": [
        {"fips": f, **v, "county": NAME[f]} for f, v in eavs.items()
    ]})
    write_json(IMPORT / "designated_county_fdic_sod.json", {"rows": [
        {"fips": f, "year": y, **vals, "county": NAME[f]}
        for f, ys in fdic.items() for y, vals in ys.items()
    ]})
    write_json(IMPORT / "designated_county_hpsa_pc.json", {"rows": [
        {"fips": f, **{k: v for k, v in vals.items()}, "county": NAME[f]}
        for f, vals in hpsa.items()
    ]})

    print(json.dumps({
        "obs": layer["stats"]["non_null_observations"],
        "added": added,
        "hypotheses": len(counter["hypotheses"]),
        "adws_families": len(adws["family_bridge"]),
    }, indent=2))


if __name__ == "__main__":
    main()
