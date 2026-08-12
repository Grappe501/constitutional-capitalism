#!/usr/bin/env python3
"""
CC-ARKANSAS-REGISTERED-VOTER-MULTIYEAR-1.0 / UPD-135
Bind EAVS 2016–2024 registered-voter series for designated AR counties into
County×Year layer + Living Systems Explorer. Observation first. 43% hold.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DL = ROOT / ".local" / "downloads" / "empirical-wave4" / "eavs"
PROJECT = ROOT / "data" / "project"
IMPORT = ROOT / "data" / "imports" / "arkansas-registered-voter-multiyear"
REPORTS = ROOT / "reports"

SLICE = "CC-ARKANSAS-REGISTERED-VOTER-MULTIYEAR-1.0"
DECISION_ID = "CC-DEC-122"
UPDATE_ID = "UPD-135"
GENERATED_AT = "2026-08-12"

FIPS = {
    "05001": "Arkansas County",
    "05073": "Lafayette County",
    "05093": "Mississippi County",
    "05107": "Phillips County",
    "05129": "Searcy County",
    "05141": "Van Buren County",
    "05145": "White County",
}

# Prefer latest corrected public releases.
YEAR_FILES = {
    "2016": {
        "kind": "csv",
        "glob": "**/EAVS_2016_Final_Data_for_Public_Release_nolabel_V1.1_CSV.csv",
        "source_id": "EAC-EAVS-2016",
        "election_type": "presidential_general_2016",
        "state_key": "State",
        "juris_key": "JurisdictionName",
        "version": "V1.1",
    },
    "2018": {
        "kind": "xlsx",
        "path": DL / "2018" / "EAVS_2018_for_Public_Release.xlsx",
        "source_id": "EAC-EAVS-2018",
        "election_type": "midterm_general_2018",
        "state_key": "State_Abbr",
        "juris_key": "Jurisdiction_Name",
        "version": "public_release",
    },
    "2020": {
        "kind": "csv",
        "glob": "**/2020_EAVS_for_Public_Release_nolabel_V1.2_CSV.csv",
        "source_id": "EAC-EAVS-2020",
        "election_type": "presidential_general_2020",
        "state_key": "State_Abbr",
        "juris_key": "Jurisdiction_Name",
        "version": "V1.2",
    },
    "2022": {
        "kind": "csv",
        "glob": "**/2022_EAVS_for_Public_Release_nolabel_V1.1_CSV.csv",
        "source_id": "EAC-EAVS-2022",
        "election_type": "midterm_general_2022",
        "state_key": "State_Abbr",
        "juris_key": "Jurisdiction_Name",
        "version": "V1.1",
    },
    "2024": {
        "kind": "csv",
        "glob": "**/2024_EAVS_for_Public_Release_nolabel_V2.csv",
        "source_id": "EAC-EAVS-2024",
        "election_type": "presidential_general_2024",
        "state_key": "State_Abbr",
        "juris_key": "Jurisdiction_Name",
        "version": "V2",
    },
}

SOURCE_URL = "https://www.eac.gov/research-and-data/studies-and-reports"


def write_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")


def num(v):
    if v in (None, "", "-99", "-88", "-77", "NA", "N/A"):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def obs(domain, metric, value, unit, definition, source_id, election_type=None, limitations=None):
    row = {
        "domain": domain,
        "metric": metric,
        "value": value,
        "unit": unit,
        "definition": definition,
        "source_id": source_id,
        "source_url": SOURCE_URL,
        "confidence": "verified_primary",
        "limitations": limitations
        or [
            "EAVS jurisdiction report for federal general election cycle",
            "Registration is a stock at report time — not identical to Election Day rolls",
        ],
    }
    if election_type:
        row["election_type"] = election_type
    return row


def iter_csv_rows(path: Path):
    with path.open(encoding="utf-8-sig", newline="", errors="replace") as f:
        yield from csv.DictReader(f)


def iter_xlsx_rows(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    for vals in rows:
        yield {header[i]: vals[i] if i < len(vals) else None for i in range(len(header))}


def load_year(year: str, meta: dict) -> dict[str, dict]:
    if meta["kind"] == "csv":
        matches = list(DL.glob(meta["glob"]))
        if not matches:
            raise FileNotFoundError(f"Missing EAVS CSV for {year}: {meta['glob']}")
        rows = iter_csv_rows(matches[0])
        file_name = matches[0].name
    else:
        path = meta["path"]
        if not path.exists():
            raise FileNotFoundError(path)
        rows = iter_xlsx_rows(path)
        file_name = path.name

    out = {}
    state_key = meta["state_key"]
    juris_key = meta["juris_key"]
    for row in rows:
        state = str(row.get(state_key) or "").strip().upper()
        if state not in ("AR", "ARKANSAS"):
            continue
        code = str(row.get("FIPSCode") or "").strip()
        # strip non-digits
        digits = "".join(ch for ch in code if ch.isdigit())
        if len(digits) < 5:
            continue
        fips = digits[:5]
        if fips not in FIPS:
            continue
        out[fips] = {
            "year": year,
            "registered_total_a1a": num(row.get("A1a")),
            "registered_active_a1b": num(row.get("A1b")),
            "registered_inactive_a1c": num(row.get("A1c")),
            "ballots_cast_f1a": num(row.get("F1a")),
            "jurisdiction": row.get(juris_key),
            "election_type": meta["election_type"],
            "source_id": meta["source_id"],
            "dataset_version": meta["version"],
            "source_file": file_name,
            "county": FIPS[fips],
        }
    if len(out) != 7:
        raise RuntimeError(f"EAVS {year}: expected 7 designated counties, got {len(out)}: {sorted(out)}")
    return out


def expand_layer(layer, series_by_year: dict[str, dict[str, dict]]) -> tuple[dict, int]:
    layer = json.loads(json.dumps(layer))
    layer["version"] = "1.3.0"
    layer["slice_id"] = "CC-ARKANSAS-COUNTY-LONGITUDINAL-OBSERVATION-LAYER-1.3"
    layer["wave_slice_id"] = SLICE
    layer["decision_id"] = DECISION_ID
    layer["update_id"] = UPDATE_ID
    layer["status"] = "LAYER_V1_3_MULTIYEAR_REGISTRATION"
    layer["generated_at"] = GENERATED_AT
    prior_obs = layer.get("stats", {}).get("non_null_observations", 0)
    county_map = {c["fips"]: c for c in layer["counties"]}
    added = 0

    def ensure_year(fips, year):
        c = county_map[fips]
        for yr in c["years"]:
            if yr["year"] == year:
                return yr
        yr = {"year": year, "observations": []}
        c["years"].append(yr)
        c["years"].sort(key=lambda x: x["year"])
        return yr

    def has(yr, metric, source_id):
        return any(
            o.get("metric") == metric
            and o.get("source_id") == source_id
            and o.get("value") is not None
            for o in yr["observations"]
        )

    def find_metric(yr, metric):
        for o in yr["observations"]:
            if o.get("metric") == metric and o.get("value") is not None:
                return o
        return None

    for year, by_fips in series_by_year.items():
        meta = YEAR_FILES[year]
        for fips, vals in by_fips.items():
            yr = ensure_year(fips, year)
            sid = vals["source_id"]
            et = vals["election_type"]

            if vals["registered_total_a1a"] is not None and not has(yr, "registered_voters_total", sid):
                # If an older same-year bind exists under a different source_id, keep both only if values differ.
                existing = find_metric(yr, "registered_voters_total")
                if existing and existing.get("source_id", "").startswith("EAC-EAVS") and existing["value"] == vals["registered_total_a1a"]:
                    pass
                else:
                    yr["observations"].append(
                        obs(
                            "civic",
                            "registered_voters_total",
                            vals["registered_total_a1a"],
                            "persons",
                            f"EAVS A1a total registered voters ({year} cycle jurisdiction report)",
                            sid,
                            election_type=et,
                            limitations=[
                                f"EAVS {year} {meta['version']} public release",
                                "Point-in-time registration stock for federal general cycle",
                                "Not SOS daily rolls; not municipal/school-board registration",
                            ],
                        )
                    )
                    added += 1

            if vals["registered_active_a1b"] is not None and not has(yr, "registered_voters_active", sid):
                yr["observations"].append(
                    obs(
                        "civic",
                        "registered_voters_active",
                        vals["registered_active_a1b"],
                        "persons",
                        f"EAVS A1b active registered voters ({year})",
                        sid,
                        election_type=et,
                    )
                )
                added += 1

            if vals["ballots_cast_f1a"] is not None and not has(yr, "ballots_cast_eavs", sid):
                yr["observations"].append(
                    obs(
                        "civic",
                        "ballots_cast_eavs",
                        vals["ballots_cast_f1a"],
                        "ballots",
                        f"EAVS F1a total ballots cast ({year} general cycle)",
                        sid,
                        election_type=et,
                    )
                )
                added += 1

            # General turnout of registered (EAVS ballots / A1a)
            if (
                vals["ballots_cast_f1a"] is not None
                and vals["registered_total_a1a"]
                and vals["registered_total_a1a"] > 0
            ):
                derived_sid = f"DERIVED-EAVS-TURNOUT-{year}"
                if not has(yr, "general_turnout_of_registered_percent", derived_sid) and not has(
                    yr, "general_turnout_of_registered_percent", "DERIVED-EAVS-TURNOUT"
                ):
                    pct = round(100.0 * vals["ballots_cast_f1a"] / vals["registered_total_a1a"], 1)
                    yr["observations"].append(
                        obs(
                            "civic",
                            "general_turnout_of_registered_percent",
                            pct,
                            "percent",
                            f"EAVS ballots cast / total registered × 100 ({year} general)",
                            derived_sid if year != "2024" else "DERIVED-EAVS-TURNOUT",
                            election_type=et,
                            limitations=[
                                "Numerator and denominator both from EAVS same-cycle report",
                                "Election-type labeled; not interchangeable with CVAP turnout",
                            ],
                        )
                    )
                    added += 1

            # Presidential turnout of registered when presidential votes exist in-year
            if year in ("2016", "2020", "2024") and vals["registered_total_a1a"]:
                pres = find_metric(yr, "presidential_total_votes")
                if pres:
                    derived_sid = f"DERIVED-PRES-EAVS-{year}"
                    # 2024 may already exist as DERIVED-PRES-EAVS
                    if not has(yr, "presidential_turnout_of_registered_percent", "DERIVED-PRES-EAVS") and not has(
                        yr, "presidential_turnout_of_registered_percent", derived_sid
                    ):
                        pct = round(100.0 * float(pres["value"]) / vals["registered_total_a1a"], 1)
                        yr["observations"].append(
                            obs(
                                "civic",
                                "presidential_turnout_of_registered_percent",
                                pct,
                                "percent",
                                f"Presidential total votes / EAVS registered total × 100 ({year})",
                                "DERIVED-PRES-EAVS" if year == "2024" else derived_sid,
                                election_type=et,
                                limitations=[
                                    f"Uses {year} EAVS registration stock with {year} presidential votes",
                                    "Votes source may be secondary canvass compilation; registration is EAVS",
                                ],
                            )
                        )
                        added += 1

    # recount non-null
    non_null = 0
    for c in layer["counties"]:
        for yr in c["years"]:
            for o in yr["observations"]:
                if o.get("value") is not None:
                    non_null += 1
    layer["stats"] = {
        "county_count": 7,
        "year_span": [
            min(yr["year"] for c in layer["counties"] for yr in c["years"]),
            max(yr["year"] for c in layer["counties"] for yr in c["years"]),
        ],
        "non_null_observations": non_null,
        "observations_added_this_wave": added,
        "prior_non_null_observations": prior_obs,
    }
    # Update NEE / blockers — multi-year registration no longer primary NEE
    nee = [d for d in layer.get("nee_domains", []) if d != "registered_voters"]
    layer["nee_domains"] = nee
    blockers = [
        b
        for b in layer.get("blockers", [])
        if b.get("id") != "registered_voter_multi_year_nee"
    ]
    blockers.insert(
        0,
        {
            "id": "registered_voter_sos_daily_rolls_nee",
            "detail": "EAVS 2016–2024 multi-year bound; Arkansas SOS daily/historical PDF rolls still NEE for intra-cycle precision",
        },
    )
    layer["blockers"] = blockers
    layer["registration_series_status"] = "EAVS_2016_2018_2020_2022_2024_BOUND"
    return layer, added


def rebuild_explorer(layer) -> dict:
    prior = json.loads((PROJECT / "arkansas_county_living_systems_explorer.json").read_text(encoding="utf-8"))
    counties = []
    for c in layer["counties"]:
        timeline = []
        for yr in c["years"]:
            metrics = {
                o["metric"]: {
                    "value": o["value"],
                    "unit": o.get("unit"),
                    "source_id": o.get("source_id"),
                    "definition": o.get("definition"),
                    "election_type": o.get("election_type"),
                }
                for o in yr["observations"]
                if o.get("value") is not None
            }
            if metrics:
                timeline.append({"year": yr["year"], "metrics": metrics})
        counties.append(
            {
                "fips": c["fips"],
                "county": c["county"],
                "role": c.get("role"),
                "timeline": timeline,
            }
        )
    prior["version"] = "1.2.0"
    prior["slice_id"] = "CC-ARKANSAS-COUNTY-LIVING-SYSTEMS-EXPLORER-1.2"
    prior["wave_slice_id"] = SLICE
    prior["decision_id"] = DECISION_ID
    prior["update_id"] = UPDATE_ID
    prior["generated_at"] = GENERATED_AT
    prior["status"] = "INTERNAL_INTEGRATED_LAB_V1_MULTIYEAR_REG"
    prior["counties"] = counties
    prior["registration_series"] = {
        "status": "EAVS_2016_2018_2020_2022_2024_BOUND",
        "years": ["2016", "2018", "2020", "2022", "2024"],
        "source_family": "EAC-EAVS",
        "note": "Federal general cycle registration stocks; SOS daily rolls still NEE",
    }
    return prior


def main():
    print("loading multi-year EAVS...")
    series = {}
    for year, meta in YEAR_FILES.items():
        series[year] = load_year(year, meta)
        print(f"  {year}: {len(series[year])} counties")

    layer = json.loads(
        (PROJECT / "arkansas_county_longitudinal_observation_layer.json").read_text(encoding="utf-8")
    )
    layer, added = expand_layer(layer, series)
    print(f"layer added observations: {added}; total non-null: {layer['stats']['non_null_observations']}")

    explorer = rebuild_explorer(layer)

    # Compact bind for imports (committed)
    bind_rows = []
    for year in sorted(series):
        for fips, vals in sorted(series[year].items()):
            bind_rows.append(vals)

    wave = {
        "version": "1.0.0",
        "slice_id": SLICE,
        "generated_at": GENERATED_AT,
        "decision_id": DECISION_ID,
        "update_id": UPDATE_ID,
        "status": "PASSED_EAVS_2016_2024_DESIGNATED_SET",
        "purpose": "Replace single-year EAVS registration snapshot with a multi-year County×Year registered-voter series for designated Arkansas counties.",
        "method_wall": [
            "observation_first",
            "interpretation_second",
            "causation_only_after_modeling",
        ],
        "not": [
            "sos_daily_roll_replacement",
            "causal_turnout_model",
            "locked_pilot_site",
            "dial_inflation",
        ],
        "years_bound": ["2016", "2018", "2020", "2022", "2024"],
        "election_types": {
            "2016": "presidential_general_2016",
            "2018": "midterm_general_2018",
            "2020": "presidential_general_2020",
            "2022": "midterm_general_2022",
            "2024": "presidential_general_2024",
        },
        "counties": 7,
        "observations_added": added,
        "layer_non_null_observations": layer["stats"]["non_null_observations"],
        "sources": [
            {"year": y, "source_id": YEAR_FILES[y]["source_id"], "version": YEAR_FILES[y]["version"]}
            for y in YEAR_FILES
        ],
        "still_nee": [
            "Arkansas SOS historical daily/PDF registration rolls",
            "Municipal / school-board registration series",
            "Primary election registration/turnout denominators",
        ],
        "holds": [
            "overall_percent_held_at_43",
            "no_locked_site",
            "election_type_labels_preserved",
            "cvap_turnout_not_confused_with_registered_turnout",
        ],
    }

    write_json(PROJECT / "arkansas_county_longitudinal_observation_layer.json", layer)
    write_json(PROJECT / "arkansas_county_living_systems_explorer.json", explorer)
    write_json(PROJECT / "arkansas_registered_voter_multiyear_wave.json", wave)
    write_json(IMPORT / "designated_county_eavs_multiyear.json", {"rows": bind_rows, "slice_id": SLICE})

    # Missing-data registry update
    gap = json.loads((PROJECT / "arkansas_county_missing_data_registry.json").read_text(encoding="utf-8"))
    for g in gap.get("gaps", []):
        if g.get("id") == "GAP-REG-MULTIYEAR":
            g["status"] = "PASSED_EAVS_2016_2024"
            g["done"] = "EAVS A1a/A1b (+F1a) for 2016/2018/2020/2022/2024 designated counties"
            g["missing"] = "Arkansas SOS daily/historical PDF rolls; municipal/school-board registration"
            g["blocks"] = ["intra-cycle registration precision", "non-federal election civic series"]
    gap["empirical_layer_checklist"]["registration"] = "PASSED_EAVS_2016_2024"
    gap["update_id"] = UPDATE_ID
    gap["decision_id"] = DECISION_ID
    gap["last_wave"] = SLICE
    write_json(PROJECT / "arkansas_county_missing_data_registry.json", gap)

    # Turnout inventory note
    inv = json.loads((PROJECT / "arkansas_county_turnout_source_inventory.json").read_text(encoding="utf-8"))
    inv["registered_voter_series_status"] = "PASSED_EAVS_2016_2024_DESIGNATED_SET"
    inv["registered_voter_series_file"] = "data/imports/arkansas-registered-voter-multiyear/designated_county_eavs_multiyear.json"
    for s in inv.get("sources", []):
        if s.get("id") == "AR-SOS-ELECTION-RESEARCH":
            s["status"] = "STILL_NEEDED_FOR_SOS_ROLLS_AND_LOCALS"
            s["note"] = "EAVS multi-year now bound; SOS still needed for daily rolls and non-federal elections"
    # ensure EAVS multi-year source entry
    if not any(s.get("id") == "EAC-EAVS-MULTIYEAR" for s in inv.get("sources", [])):
        inv["sources"].insert(
            0,
            {
                "id": "EAC-EAVS-MULTIYEAR",
                "name": "EAC Election Administration and Voting Survey (2016–2024)",
                "url": SOURCE_URL,
                "coverage": "Designated AR counties: A1a registration (+A1b where available), F1a ballots; presidential + midterm generals",
                "status": "BOUND_FOR_DESIGNATED_COUNTIES",
                "years_bound": ["2016", "2018", "2020", "2022", "2024"],
            },
        )
    inv["update_id_registration"] = UPDATE_ID
    write_json(PROJECT / "arkansas_county_turnout_source_inventory.json", inv)

    print("done", SLICE, "added", added)


if __name__ == "__main__":
    main()
